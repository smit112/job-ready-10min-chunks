"""
Excel processor for configuration templates and validation rules.
Handles reading, writing, and processing Excel files for configuration research.
"""
import openpyxl
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
import logging
from dataclasses import dataclass
import json

logger = logging.getLogger(__name__)


@dataclass
class ExcelConfig:
    """Configuration data extracted from Excel files."""
    sheet_name: str
    headers: List[str]
    data: List[Dict[str, Any]]
    validation_rules: Dict[str, Any]
    metadata: Dict[str, Any]


class ExcelProcessor:
    """Processes Excel files for configuration research and validation."""
    
    def __init__(self, templates_dir: str, output_dir: str):
        self.templates_dir = Path(templates_dir)
        self.output_dir = Path(output_dir)
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def read_excel_file(self, file_path: Union[str, Path]) -> Dict[str, ExcelConfig]:
        """
        Read Excel file and extract configuration data from all sheets.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary mapping sheet names to ExcelConfig objects
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        configs = {}
        
        try:
            # Read Excel file using openpyxl
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Extract headers from first row
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                    else:
                        headers.append(f"Column_{len(headers) + 1}")
                
                # Extract data from remaining rows
                data = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        row_data = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = value
                        data.append(row_data)
                
                # Extract validation rules (simplified)
                validation_rules = self._extract_validation_rules_simple(worksheet, headers)
                
                # Extract metadata
                metadata = {
                    'file_name': file_path.name,
                    'sheet_name': sheet_name,
                    'row_count': len(data),
                    'column_count': len(headers),
                    'last_modified': file_path.stat().st_mtime
                }
                
                configs[sheet_name] = ExcelConfig(
                    sheet_name=sheet_name,
                    headers=headers,
                    data=data,
                    validation_rules=validation_rules,
                    metadata=metadata
                )
                
                logger.info(f"Processed sheet '{sheet_name}' with {len(data)} rows")
        
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            raise
        
        return configs
    
    def _extract_validation_rules_simple(self, worksheet, headers: List[str]) -> Dict[str, Any]:
        """Extract validation rules from worksheet."""
        rules = {}
        
        for i, header in enumerate(headers):
            # Check for validation rule indicators in column names
            if 'validation' in header.lower() or 'rule' in header.lower():
                # Get values from this column
                values = []
                for row in worksheet.iter_rows(min_row=2, min_col=i+1, max_col=i+1, values_only=True):
                    if row[0] is not None:
                        values.append(str(row[0]))
                rules[header] = values
            elif 'required' in header.lower():
                rules[f"{header}_required"] = True
            elif 'format' in header.lower():
                # Get format values from this column
                values = []
                for row in worksheet.iter_rows(min_row=2, min_col=i+1, max_col=i+1, values_only=True):
                    if row[0] is not None:
                        values.append(str(row[0]))
                rules[f"{header}_format"] = values
        
        return rules
    
    def create_configuration_template(self, 
                                    template_name: str,
                                    config_schema: Dict[str, Any],
                                    output_file: Optional[str] = None) -> Path:
        """
        Create a configuration template Excel file.
        
        Args:
            template_name: Name of the template
            config_schema: Schema defining the configuration structure
            output_file: Optional output file path
            
        Returns:
            Path to the created template file
        """
        if output_file is None:
            output_file = self.templates_dir / f"{template_name}_template.xlsx"
        else:
            output_file = Path(output_file)
        
        # Create workbook and worksheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create configuration sheet
        config_sheet = wb.create_sheet("Configuration")
        validation_sheet = wb.create_sheet("Validation_Rules")
        metadata_sheet = wb.create_sheet("Metadata")
        
        # Populate configuration sheet
        row = 1
        for field_name, field_info in config_schema.items():
            config_sheet.cell(row=row, column=1, value=field_name)
            config_sheet.cell(row=row, column=2, value=field_info.get('description', ''))
            config_sheet.cell(row=row, column=3, value=field_info.get('type', 'string'))
            config_sheet.cell(row=row, column=4, value=field_info.get('default', ''))
            config_sheet.cell(row=row, column=5, value=field_info.get('required', False))
            row += 1
        
        # Add headers
        config_sheet.cell(row=1, column=1, value="Field Name")
        config_sheet.cell(row=1, column=2, value="Description")
        config_sheet.cell(row=1, column=3, value="Type")
        config_sheet.cell(row=1, column=4, value="Default Value")
        config_sheet.cell(row=1, column=5, value="Required")
        
        # Populate validation rules sheet
        validation_sheet.cell(row=1, column=1, value="Field Name")
        validation_sheet.cell(row=1, column=2, value="Validation Rule")
        validation_sheet.cell(row=1, column=3, value="Error Message")
        
        row = 2
        for field_name, field_info in config_schema.items():
            if 'validation' in field_info:
                validation_sheet.cell(row=row, column=1, value=field_name)
                validation_sheet.cell(row=row, column=2, value=field_info['validation'])
                validation_sheet.cell(row=row, column=3, value=field_info.get('error_message', ''))
                row += 1
        
        # Populate metadata sheet
        metadata_sheet.cell(row=1, column=1, value="Property")
        metadata_sheet.cell(row=1, column=2, value="Value")
        
        from datetime import datetime
        metadata = {
            "Template Name": template_name,
            "Created": datetime.now().isoformat(),
            "Version": "1.0",
            "Description": f"Configuration template for {template_name}"
        }
        
        row = 2
        for prop, value in metadata.items():
            metadata_sheet.cell(row=row, column=1, value=prop)
            metadata_sheet.cell(row=row, column=2, value=value)
            row += 1
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Created configuration template: {output_file}")
        
        return output_file
    
    def validate_configuration(self, 
                             config_data: Dict[str, Any],
                             validation_rules: Dict[str, Any]) -> Dict[str, List[str]]:
        """
        Validate configuration data against validation rules.
        
        Args:
            config_data: Configuration data to validate
            validation_rules: Validation rules to apply
            
        Returns:
            Dictionary of validation errors by field
        """
        errors = {}
        
        for field_name, rules in validation_rules.items():
            field_errors = []
            field_value = config_data.get(field_name)
            
            # Check required fields
            if rules.get('required', False) and (field_value is None or field_value == ''):
                field_errors.append(f"Field '{field_name}' is required")
            
            # Check data type
            expected_type = rules.get('type', 'string')
            if field_value is not None and not self._check_type(field_value, expected_type):
                field_errors.append(f"Field '{field_name}' must be of type {expected_type}")
            
            # Check format
            if 'format' in rules and field_value is not None:
                if not self._check_format(field_value, rules['format']):
                    field_errors.append(f"Field '{field_name}' format is invalid")
            
            # Check range
            if 'min' in rules and field_value is not None:
                if field_value < rules['min']:
                    field_errors.append(f"Field '{field_name}' must be >= {rules['min']}")
            
            if 'max' in rules and field_value is not None:
                if field_value > rules['max']:
                    field_errors.append(f"Field '{field_name}' must be <= {rules['max']}")
            
            if field_errors:
                errors[field_name] = field_errors
        
        return errors
    
    def _check_type(self, value: Any, expected_type: str) -> bool:
        """Check if value matches expected type."""
        type_mapping = {
            'string': str,
            'integer': int,
            'float': float,
            'boolean': bool,
            'list': list,
            'dict': dict
        }
        
        expected_python_type = type_mapping.get(expected_type, str)
        return isinstance(value, expected_python_type)
    
    def _check_format(self, value: Any, format_pattern: str) -> bool:
        """Check if value matches format pattern."""
        import re
        try:
            return bool(re.match(format_pattern, str(value)))
        except re.error:
            logger.warning(f"Invalid regex pattern: {format_pattern}")
            return False
    
    def export_validation_results(self, 
                                validation_results: Dict[str, List[str]],
                                output_file: Optional[str] = None) -> Path:
        """
        Export validation results to Excel file.
        
        Args:
            validation_results: Validation results to export
            output_file: Optional output file path
            
        Returns:
            Path to the exported file
        """
        if output_file is None:
            from datetime import datetime
            output_file = self.output_dir / f"validation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            output_file = Path(output_file)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create validation results sheet
        results_sheet = wb.create_sheet("Validation_Results")
        
        # Add headers
        results_sheet.cell(row=1, column=1, value="Field Name")
        results_sheet.cell(row=1, column=2, value="Error")
        results_sheet.cell(row=1, column=3, value="Status")
        results_sheet.cell(row=1, column=4, value="Timestamp")
        
        # Add data
        row = 2
        for field_name, errors in validation_results.items():
            for error in errors:
                results_sheet.cell(row=row, column=1, value=field_name)
                results_sheet.cell(row=row, column=2, value=error)
                results_sheet.cell(row=row, column=3, value="FAILED")
                results_sheet.cell(row=row, column=4, value=datetime.now().isoformat())
                row += 1
        
        if not any(validation_results.values()):
            results_sheet.cell(row=2, column=1, value="All Fields")
            results_sheet.cell(row=2, column=2, value="No validation errors found")
            results_sheet.cell(row=2, column=3, value="PASSED")
            results_sheet.cell(row=2, column=4, value=datetime.now().isoformat())
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.cell(row=1, column=1, value="Property")
        summary_sheet.cell(row=1, column=2, value="Value")
        
        summary_data = {
            'Total Fields': len(validation_results),
            'Fields with Errors': len([f for f, e in validation_results.items() if e]),
            'Total Errors': sum(len(errors) for errors in validation_results.values()),
            'Validation Date': datetime.now().isoformat()
        }
        
        row = 2
        for prop, value in summary_data.items():
            summary_sheet.cell(row=row, column=1, value=prop)
            summary_sheet.cell(row=row, column=2, value=value)
            row += 1
        
        # Save workbook
        wb.save(output_file)
        
        logger.info(f"Exported validation results to: {output_file}")
        return output_file